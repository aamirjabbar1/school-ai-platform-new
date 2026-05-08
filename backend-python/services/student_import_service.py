"""
Student Import Service
======================
Parses student list PDF, extracts student names,
creates accounts in bulk, and generates credentials file.
"""

import re
import secrets
import string
from io import BytesIO
import pdfplumber
from passlib.hash import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models.models import User


# ─── PDF PARSING ──────────────────────────────────────────────────────────────

# Column boundaries (from PDF header x-coordinates)
STUDENT_X_START = 190
FATHER_X_START = 335


def parse_student_pdf(file_bytes: bytes) -> list[dict]:
    """
    Parse student list PDF and extract student records.
    Returns list of dicts: {name, class_section}
    """
    pdf = pdfplumber.open(BytesIO(file_bytes))

    students = []
    seen = set()
    current_class = ""

    for page in pdf.pages:
        words = page.extract_words(x_tolerance=2, y_tolerance=2)

        # Group words by row
        rows = {}
        for w in words:
            row_key = round(w["top"] * 2) / 2
            if row_key not in rows:
                rows[row_key] = []
            rows[row_key].append(w)

        for row_key in sorted(rows.keys()):
            row_words = sorted(rows[row_key], key=lambda x: x["x0"])
            full_text = " ".join(w["text"] for w in row_words)

            # Skip header rows
            if "Student" in full_text and "Father" in full_text:
                continue

            has_gender = any(w["text"] in ("Male", "Female") for w in row_words)

            # Detect class/section headers
            if not has_gender:
                # Class headers like "Pre-Nursery - Disney", "One - Planes", "Ten - -X-B"
                header_keywords = [
                    "Nursery", "KG", "One", "Two", "Three", "Four", "Five",
                    "Six", "Seven", "Eight", "Nine", "Ten", "Pre-Nine", "Prep",
                ]
                if any(kw in full_text for kw in header_keywords):
                    # Clean up: skip if it looks like a data row (has LSS reg number)
                    if "LSS" not in full_text:
                        current_class = _normalize_class(full_text.strip())
                continue

            # Extract student name from the Student Name column
            student_words = []
            for w in row_words:
                center = (w["x0"] + w["x1"]) / 2
                # Word center is in student name column
                if STUDENT_X_START <= center <= FATHER_X_START:
                    cleaned = re.sub(r"^\d+[-_]?\d*", "", w["text"])
                    if cleaned and not cleaned.isdigit():
                        student_words.append(cleaned)
                # Word starts in GR column but extends into student column
                elif w["x0"] < STUDENT_X_START and w["x1"] > STUDENT_X_START + 10:
                    cleaned = re.sub(r"^\d+[-_]?\d*", "", w["text"])
                    if cleaned and len(cleaned) > 1 and not cleaned.isdigit():
                        student_words.append(cleaned)

            student_name = " ".join(student_words).strip()

            if not student_name or len(student_name) < 3:
                continue

            # Deduplicate: allow same name in different classes
            dedup_key = f"{student_name.upper()}|{current_class}"
            if dedup_key in seen:
                continue
            seen.add(dedup_key)

            students.append({
                "name": student_name.title(),
                "class_section": current_class,
            })

    pdf.close()
    return students


def _normalize_class(raw: str) -> str:
    """Normalize class name like 'One - Planes' -> 'Class 1 - Planes'"""
    mapping = {
        "Pre-Nursery": "Pre-Nursery",
        "Nursery": "Nursery",
        "KG": "KG",
        "One": "Class 1",
        "Two": "Class 2",
        "Three": "Class 3",
        "Four": "Class 4",
        "Five": "Class 5",
        "Six": "Class 6",
        "Seven": "Class 7",
        "Eight": "Class 8",
        "Pre-Nine": "Class 8",
        "Nine": "Class 9",
        "Ten": "Class 10",
    }
    for key, val in mapping.items():
        if raw.startswith(key + " "):
            rest = raw[len(key):].strip()
            # Clean section names like "- -III-B-I" -> "Boys", "- -III-G-I" -> "Girls"
            if "-B-" in rest or "(B)" in rest:
                section = "Boys"
            elif "-G-" in rest or "(G)" in rest:
                section = "Girls"
            else:
                section = rest.lstrip("- ").strip()
            return f"{val} - {section}" if section else val
        elif raw == key:
            return val
    return raw


# ─── PASSWORD GENERATION ──────────────────────────────────────────────────────

def generate_password(length: int = 8) -> str:
    """Generate a random password (simpler for students/parents)."""
    chars = string.ascii_letters + string.digits
    password = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
    ]
    password += [secrets.choice(chars) for _ in range(length - 3)]
    pw_list = list(password)
    secrets.SystemRandom().shuffle(pw_list)
    return "".join(pw_list)


def generate_login_id(name: str, counter: int, existing_ids: set) -> str:
    """
    Generate a unique student login ID.
    Format: STU + 4-digit number (e.g. STU0001)
    """
    login_id = f"STU{counter:04d}"
    while login_id.lower() in existing_ids:
        counter += 1
        login_id = f"STU{counter:04d}"
    return login_id


# ─── ACCOUNT CREATION ────────────────────────────────────────────────────────

async def create_student_accounts(
    records: list[dict],
    db: AsyncSession,
) -> dict:
    """
    Create student accounts for all parsed records.
    Returns: {created: [...], skipped: [...], total: int}
    """
    # Get existing login IDs
    result = await db.execute(select(User.login_id))
    existing_login_ids = {r[0].lower() for r in result.all()}

    # Get existing student names
    result2 = await db.execute(select(User.name).where(User.role == "student"))
    existing_names = {r[0].lower() for r in result2.all()}

    # Find max existing STU number
    max_stu = 0
    for lid in existing_login_ids:
        m = re.match(r"stu(\d+)", lid)
        if m:
            max_stu = max(max_stu, int(m.group(1)))

    counter = max_stu + 1
    created = []
    skipped = []

    for rec in records:
        name_lower = rec["name"].lower()

        # Skip if exact name already exists
        if name_lower in existing_names:
            skipped.append({**rec, "reason": "Name already exists"})
            continue

        login_id = generate_login_id(rec["name"], counter, existing_login_ids)
        counter = int(login_id[3:]) + 1
        password = generate_password()

        # Extract class level for user record
        class_name = rec["class_section"].split(" - ")[0] if " - " in rec["class_section"] else rec["class_section"]

        new_user = User(
            name=rec["name"],
            login_id=login_id,
            password_hash=bcrypt.hash(password),
            role="student",
            class_name=class_name,
            is_active=True,
        )
        db.add(new_user)
        existing_login_ids.add(login_id.lower())
        existing_names.add(name_lower)

        created.append({
            **rec,
            "login_id": login_id,
            "password": password,
        })

    await db.commit()

    return {
        "created": created,
        "skipped": skipped,
        "total": len(records),
    }


# ─── EXCEL GENERATION ────────────────────────────────────────────────────────

def generate_student_credentials_excel(created: list[dict], skipped: list[dict]) -> bytes:
    """Generate Excel file with student credentials grouped by class."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Student Credentials"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    class_font = Font(name="Calibri", bold=True, size=11, color="C55A11")
    class_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "LSS Bot - Student Login Credentials"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = "CONFIDENTIAL - Share with parents only"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="FF0000")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Group by class
    from collections import OrderedDict
    by_class = OrderedDict()
    for rec in created:
        cls = rec.get("class_section", "Unknown")
        if cls not in by_class:
            by_class[cls] = []
        by_class[cls].append(rec)

    row = 4
    student_num = 0
    for cls_name, cls_students in by_class.items():
        # Class header
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = cls_name
        ws[f"A{row}"].font = class_font
        ws[f"A{row}"].fill = class_fill
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = class_fill
            ws.cell(row=row, column=col).border = border
        row += 1

        # Column headers
        headers = ["#", "Student Name", "Class", "User ID", "Password"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        row += 1

        # Student rows
        for rec in cls_students:
            student_num += 1
            ws.cell(row=row, column=1, value=student_num).border = border
            ws.cell(row=row, column=2, value=rec["name"]).border = border
            ws.cell(row=row, column=3, value=rec.get("class_section", "")).border = border
            ws.cell(row=row, column=4, value=rec["login_id"]).border = border
            ws.cell(row=row, column=5, value=rec["password"]).border = border

            if student_num % 2 == 0:
                light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = light_fill

            row += 1

        row += 1  # Gap between classes

    # Summary
    ws.cell(row=row, column=1, value=f"Total Students: {len(created)}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    # Skipped sheet
    if skipped:
        ws2 = wb.create_sheet("Skipped")
        skip_headers = ["#", "Name", "Class", "Reason"]
        for col, h in enumerate(skip_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.border = border
        for i, rec in enumerate(skipped, 1):
            ws2.cell(row=i + 1, column=1, value=i).border = border
            ws2.cell(row=i + 1, column=2, value=rec["name"]).border = border
            ws2.cell(row=i + 1, column=3, value=rec.get("class_section", "")).border = border
            ws2.cell(row=i + 1, column=4, value=rec.get("reason", "")).border = border

        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 30
        ws2.column_dimensions["C"].width = 22
        ws2.column_dimensions["D"].width = 25

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
