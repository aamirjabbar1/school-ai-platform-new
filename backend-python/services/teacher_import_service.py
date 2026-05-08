"""
Teacher Import Service
======================
Parses salary PDF from LSSP HR system, extracts teacher data,
creates accounts in bulk, and generates credentials file.
"""

import re
import secrets
import string
from io import BytesIO
from pypdf import PdfReader
from passlib.hash import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models.models import User


# ─── PDF PARSING ──────────────────────────────────────────────────────────────

def parse_salary_pdf(file_bytes: bytes) -> list[dict]:
    """
    Parse LSSP salary PDF and extract staff records.
    Returns list of dicts: {sr, reg_no, name, dept, designation}
    """
    reader = PdfReader(BytesIO(file_bytes))
    all_text = ""
    for page in reader.pages:
        text = page.extract_text()
        if text:
            all_text += text + "\n"

    records = []
    seen = set()

    lines = all_text.replace("\n", " ")

    # Two-step parse: first find reg numbers, then extract surrounding data
    # Reg numbers always match: 20xxLSS + digits (e.g. 2024LSS3582, 2021LSS13428, 2014LSS010012)
    pattern = re.compile(
        r'(20[12]\dLSS\d{3,7})'              # Registration Number
        r'\s*'
        r'([A-Z][A-Z\s\.]+?)'                # Name (uppercase letters, spaces, dots)
        r'\s*'
        r'(Teachers|Admin|Media|Management|Accounts|Sports|Teache[r]?s?)'  # Department
        r'\s*'
        r'(Teacher|Admin|Principal|Manager|Media|ACCOUNTANT|PTI)'          # Designation
        r'\s*Bank',                            # Pay mode
        re.IGNORECASE
    )

    sr_counter = 0
    for match in pattern.finditer(lines):
        sr_counter += 1
        reg_no = match.group(1)
        name = match.group(2).strip()
        dept = match.group(3).strip()
        desg = match.group(4).strip()
        sr = str(sr_counter)

        # Deduplicate by registration number
        if reg_no in seen:
            continue
        seen.add(reg_no)

        # Clean name - remove trailing single letters
        name = re.sub(r'\s+', ' ', name).strip()

        records.append({
            "sr": int(sr),
            "reg_no": reg_no,
            "name": name,
            "dept": dept,
            "designation": desg,
        })

    # Sort by sr number
    records.sort(key=lambda r: r["sr"])
    return records


# ─── PASSWORD GENERATION ──────────────────────────────────────────────────────

def generate_password(length: int = 10) -> str:
    """Generate a strong random password."""
    chars = string.ascii_letters + string.digits + "!@#$%"
    # Ensure at least one of each type
    password = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
        secrets.choice("!@#$%"),
    ]
    password += [secrets.choice(chars) for _ in range(length - 4)]
    # Shuffle to avoid predictable positions
    pw_list = list(password)
    secrets.SystemRandom().shuffle(pw_list)
    return "".join(pw_list)


def generate_login_id(name: str, reg_no: str, existing_ids: set) -> str:
    """
    Generate a unique login ID from name.
    Format: first letter of each name part + last 4 digits of reg_no
    e.g., AFSHEEN WALI -> AW3582
    """
    parts = name.split()
    initials = "".join(p[0].upper() for p in parts if p)
    # Get last 4-5 digits from reg_no
    digits = re.findall(r'\d+', reg_no)
    suffix = digits[-1][-4:] if digits else reg_no[-4:]

    login_id = f"{initials}{suffix}"

    # Ensure uniqueness
    base = login_id
    counter = 1
    while login_id.lower() in existing_ids:
        login_id = f"{base}{counter}"
        counter += 1

    return login_id


# ─── ACCOUNT CREATION ────────────────────────────────────────────────────────

async def create_teacher_accounts(
    records: list[dict],
    db: AsyncSession,
) -> dict:
    """
    Create user accounts for all parsed records.
    Returns: {created: [...], skipped: [...], total: int}
    """
    # Get all existing login IDs and reg numbers
    result = await db.execute(select(User.login_id))
    existing_login_ids = {r[0].lower() for r in result.all()}

    # Also check by name to avoid double-creating
    result2 = await db.execute(select(User.name))
    existing_names = {r[0].lower() for r in result2.all()}

    created = []
    skipped = []

    for rec in records:
        # Skip if name already exists (case-insensitive)
        if rec["name"].lower() in existing_names:
            skipped.append({
                **rec,
                "reason": "Name already exists in system",
            })
            continue

        # Generate credentials
        login_id = generate_login_id(rec["name"], rec["reg_no"], existing_login_ids)
        password = generate_password()

        # Determine role based on department/designation
        role = "teacher"
        if rec["designation"].lower() in ("admin", "principal", "manager", "accountant"):
            role = "admin"

        # Create user
        new_user = User(
            name=rec["name"].title(),  # Title case
            login_id=login_id,
            password_hash=bcrypt.hash(password),
            role=role,
            is_active=True,
        )
        db.add(new_user)
        existing_login_ids.add(login_id.lower())
        existing_names.add(rec["name"].lower())

        created.append({
            **rec,
            "login_id": login_id,
            "password": password,  # Plain text for the output file only
            "role": role,
        })

    await db.commit()

    return {
        "created": created,
        "skipped": skipped,
        "total": len(records),
    }


# ─── EXCEL GENERATION ────────────────────────────────────────────────────────

def generate_credentials_excel(created: list[dict], skipped: list[dict]) -> bytes:
    """
    Generate an Excel file with teacher credentials.
    Returns bytes of the .xlsx file.
    """
    wb = Workbook()

    # ── Sheet 1: Credentials ──
    ws = wb.active
    ws.title = "Teacher Credentials"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=16, color="2F5496")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "LSS Bot - Teacher Login Credentials"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = "CONFIDENTIAL - Distribute securely to each teacher"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="FF0000")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Sr.", "Name", "Registration No.", "User ID", "Password", "Role"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data rows
    for i, rec in enumerate(created, 1):
        row = i + 4
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=rec["name"].title()).border = border
        ws.cell(row=row, column=3, value=rec["reg_no"]).border = border
        ws.cell(row=row, column=4, value=rec["login_id"]).border = border
        ws.cell(row=row, column=5, value=rec["password"]).border = border
        ws.cell(row=row, column=6, value=rec["role"].title()).border = border

        # Alternate row shading
        if i % 2 == 0:
            light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = light_fill

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10

    # Summary row
    summary_row = len(created) + 6
    ws.cell(row=summary_row, column=1, value=f"Total Accounts Created: {len(created)}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=11)

    # ── Sheet 2: Skipped (if any) ──
    if skipped:
        ws2 = wb.create_sheet("Skipped Entries")
        skip_headers = ["Sr.", "Name", "Registration No.", "Reason"]
        for col, header in enumerate(skip_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.border = border

        for i, rec in enumerate(skipped, 1):
            ws2.cell(row=i + 1, column=1, value=i).border = border
            ws2.cell(row=i + 1, column=2, value=rec["name"].title()).border = border
            ws2.cell(row=i + 1, column=3, value=rec["reg_no"]).border = border
            ws2.cell(row=i + 1, column=4, value=rec.get("reason", "")).border = border

        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 28
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 30

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
