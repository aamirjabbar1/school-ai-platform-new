"""
Bulk student import from an Excel (.xlsx) file.

Expected columns (header names matched flexibly, case/spacing-insensitive):
    Registration Number | Student Name | Father Name | Class

Pipeline (run inside a Celery task — see tasks/student_import_tasks.py):
  1. Parse rows from the workbook.
  2. Validate (registration number, name, class all required; class must resolve
     to a known class) — bad rows are reported, never abort the whole import.
  3. Handle duplicates by registration number per the chosen mode
     (skip / update / create_new).
  4. Create login accounts (login_id = registration number) + set passwords
     per the chosen password mode (registration / custom / random).
  5. Generate a credentials Excel for download.
  6. Record a full summary + the ids of created accounts (for rollback) on the
     StudentImportBatch row.

Designed to handle ~1000 records in one upload: a single query loads existing
accounts, all inserts/updates are committed once, and the work runs in the
background worker so the HTTP request returns immediately.
"""
from __future__ import annotations

import io
import re
import secrets
import string
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models.models import User, StudentImportBatch
from utils.password import hash_password


# ─── CANONICAL CLASSES + NORMALIZATION ───────────────────────────────────────

CANONICAL_CLASSES = [
    "Pre-Nursery", "Nursery", "KG",
    "Class 1", "Class 2", "Class 3", "Class 4", "Class 5", "Class 6",
    "Class 7", "Class 8", "Class 9", "Class 10", "Class 11", "Class 12",
]

_WORD_NUM = {
    "one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6,
    "seven": 7, "eight": 8, "nine": 9, "ten": 10, "eleven": 11, "twelve": 12,
    "first": 1, "second": 2, "third": 3, "fourth": 4, "fifth": 5, "sixth": 6,
    "seventh": 7, "eighth": 8, "ninth": 9, "tenth": 10, "eleventh": 11, "twelfth": 12,
}
_ROMAN = {
    "i": 1, "ii": 2, "iii": 3, "iv": 4, "v": 5, "vi": 6,
    "vii": 7, "viii": 8, "ix": 9, "x": 10, "xi": 11, "xii": 12,
}


def normalize_class(raw) -> str | None:
    """Map a free-text class value to a canonical class name, or None if it
    can't be matched to a class the system knows about.

    Handles: "Class 8", "Grade 8", "8", "VIII", "Eight", "8-A", "Class 8 Boys",
    "Pre-Nursery", "Nursery", "KG", "Prep", "Pre-Nine" (→ Class 8), etc.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None

    low = re.sub(r"[\-_]", " ", s.lower())
    low = re.sub(r"\s+", " ", low).strip()

    # Pre-primary levels
    if "pre" in low and "nursery" in low:
        return "Pre-Nursery"
    if low in ("nursery", "nur"):
        return "Nursery"
    if low in ("kg", "k g", "kindergarten", "prep", "prep class", "preparatory"):
        return "KG"
    if low.startswith("kg") or "kindergarten" in low:
        return "KG"
    # Pre-9th convention → studies the Class 8 record (matches student importer)
    if "pre nine" in low or "pre 9" in low or "pre ix" in low:
        return "Class 8"

    # Strip class/grade keywords, then look for a number (digit, word or roman)
    cleaned = re.sub(r"\b(class|grade|std|standard|level|year|section|sec)\b", " ", low)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()

    n: int | None = None
    m = re.search(r"\b(\d{1,2})\b", cleaned)
    if m:
        n = int(m.group(1))
    else:
        for tok in cleaned.split():
            if tok in _WORD_NUM:
                n = _WORD_NUM[tok]
                break
            if tok in _ROMAN:
                n = _ROMAN[tok]
                break

    if n is not None and 1 <= n <= 12:
        return f"Class {n}"
    return None


def normalize_section(raw) -> str | None:
    """Clean a section value, or None if empty. Strips a leading 'Section'/'Sec'
    keyword and upper-cases short codes ("a" → "A") while leaving named sections
    (e.g. "Blue") otherwise intact."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    s = re.sub(r"^(section|sec)\b[\s:.\-]*", "", s, flags=re.IGNORECASE).strip()
    s = re.sub(r"\s+", " ", s)
    if not s:
        return None
    return s.upper() if len(s) <= 3 else s


# ─── EXCEL PARSING ────────────────────────────────────────────────────────────

def _norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _classify_header(norm: str) -> str | None:
    """Return which logical column a header cell maps to, or None."""
    if not norm:
        return None
    if "father" in norm:
        return "father_name"
    if (
        "registration" in norm or norm.startswith("reg") or "regno" in norm
        or norm in ("grno", "gr", "admissionno", "admissionnumber", "rollno", "rollnumber")
        or "admission" in norm
    ):
        return "reg_no"
    if "studentname" in norm or norm in ("name", "studentsname", "fullname", "studentfullname") or "student" in norm:
        return "name"
    if "section" in norm:
        return "section"
    if "class" in norm or "grade" in norm:
        return "class_name"
    return None


def _cell_str(value) -> str:
    """Coerce an openpyxl cell value to a clean string. Numeric registration
    numbers come back as int/float — render those without a trailing .0."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def parse_excel(file_bytes: bytes) -> list[dict]:
    """Parse the workbook into a list of raw row dicts:
        {row, reg_no, name, father_name, class_raw}
    Raises ValueError if the file can't be read or required columns are missing.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read the Excel file: {exc}")

    ws = wb.active
    if ws is None:
        raise ValueError("The Excel file has no worksheet")

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError("The Excel file is empty")

    # Locate the header row within the first 10 rows
    col_map: dict[str, int] = {}
    header_idx = -1
    for idx, row in enumerate(rows[:10]):
        mapping: dict[str, int] = {}
        for ci, cell in enumerate(row):
            logical = _classify_header(_norm_header(cell))
            if logical and logical not in mapping:
                mapping[logical] = ci
        if "reg_no" in mapping and "name" in mapping and "class_name" in mapping:
            col_map = mapping
            header_idx = idx
            break

    if header_idx < 0:
        raise ValueError(
            "Could not find the required columns. The file must have headers "
            "for Registration Number, Student Name and Class."
        )

    parsed: list[dict] = []
    for offset, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        reg_no = _cell_str(row[col_map["reg_no"]]) if col_map["reg_no"] < len(row) else ""
        name = _cell_str(row[col_map["name"]]) if col_map["name"] < len(row) else ""
        class_raw = _cell_str(row[col_map["class_name"]]) if col_map["class_name"] < len(row) else ""
        father = ""
        if "father_name" in col_map and col_map["father_name"] < len(row):
            father = _cell_str(row[col_map["father_name"]])
        section_raw = ""
        if "section" in col_map and col_map["section"] < len(row):
            section_raw = _cell_str(row[col_map["section"]])

        # Skip wholly blank rows silently
        if not (reg_no or name or class_raw or father or section_raw):
            continue

        parsed.append({
            "row": offset,
            "reg_no": reg_no,
            "name": name,
            "father_name": father,
            "class_raw": class_raw,
            "section_raw": section_raw,
        })

    return parsed


# ─── PASSWORD GENERATION ──────────────────────────────────────────────────────

def _random_password(length: int = 8) -> str:
    chars = string.ascii_letters + string.digits
    pw = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
    ] + [secrets.choice(chars) for _ in range(length - 3)]
    secrets.SystemRandom().shuffle(pw)
    return "".join(pw)


# ─── MAIN PROCESSING ──────────────────────────────────────────────────────────

def _utcnow():
    return datetime.now(timezone.utc).replace(tzinfo=None)


async def process_student_import(
    batch_id: str,
    file_bytes: bytes,
    db: AsyncSession,
    custom_password: str | None = None,
) -> None:
    """Run the full import for a batch row. Updates the batch in place with the
    final status, summary counts, error log, created account ids and credentials.
    Never raises — a fatal error is recorded on the batch as status='failed'.

    `custom_password` is only used when password_mode == 'custom'. It is passed
    through from the request and never persisted to the database.
    """
    result = await db.execute(select(StudentImportBatch).where(StudentImportBatch.id == batch_id))
    batch = result.scalar_one_or_none()
    if not batch:
        return

    batch.status = "processing"
    await db.commit()

    try:
        rows = parse_excel(file_bytes)
        batch.total = len(rows)
        await db.commit()

        dup_mode = batch.duplicate_mode or "skip"
        pw_mode = batch.password_mode or "registration"
        section_mode = batch.section_mode or "create"
        custom_pw = (custom_password or "").strip()

        # Pre-load existing accounts by login_id (registration number) in one query.
        reg_numbers = [r["reg_no"].strip() for r in rows if r["reg_no"].strip()]
        existing_by_login: dict[str, User] = {}
        if reg_numbers:
            res = await db.execute(select(User).where(User.login_id.in_(reg_numbers)))
            for u in res.scalars().all():
                existing_by_login[u.login_id] = u

        # For strict section mode, pre-load the class+section combinations that
        # already exist among students — a section is "known" only if some student
        # already belongs to it. (key = (class_name, section.lower()))
        existing_combos: set[tuple[str, str]] = set()
        if section_mode == "strict":
            res = await db.execute(
                select(User.class_name, User.section).where(
                    User.role == "student", User.section.isnot(None)
                )
            )
            for cn, sec in res.all():
                if cn and sec:
                    existing_combos.add((cn, str(sec).strip().lower()))

        # Pre-compute the shared hash once for custom-password mode (huge speedup).
        custom_hash = hash_password(custom_pw) if (pw_mode == "custom" and custom_pw) else None

        created_count = updated_count = skipped_count = failed_count = 0
        error_log: list[dict] = []
        created_ids: list[str] = []
        credentials_rows: list[dict] = []
        seen_regs: set[str] = set()
        new_users: list[User] = []

        for r in rows:
            reg = r["reg_no"].strip()
            name = r["name"].strip()
            father = r["father_name"].strip()
            class_raw = r["class_raw"].strip()
            section_raw = r.get("section_raw", "").strip()

            # ── Validation ──
            missing = []
            if not reg:
                missing.append("Registration Number")
            if not name:
                missing.append("Student Name")
            if not class_raw:
                missing.append("Class")
            if not section_raw:
                missing.append("Section")
            if missing:
                failed_count += 1
                error_log.append({"row": r["row"], "reg_no": reg, "name": name,
                                  "reason": f"Missing required field(s): {', '.join(missing)}"})
                continue

            # Registration number is used as the login id (max 50 chars)
            if len(reg) > 50:
                failed_count += 1
                error_log.append({"row": r["row"], "reg_no": reg[:50], "name": name,
                                  "reason": "Registration number is too long (max 50 characters)"})
                continue

            # Duplicate registration number within the same file
            if reg in seen_regs:
                failed_count += 1
                error_log.append({"row": r["row"], "reg_no": reg, "name": name,
                                  "reason": "Duplicate registration number in file"})
                continue
            seen_regs.add(reg)

            # Defensive truncation so one oversized cell can't fail the batch
            name = name[:100]
            father = father[:150]

            # Class must resolve to a known class
            class_name = normalize_class(class_raw)
            if not class_name:
                failed_count += 1
                error_log.append({"row": r["row"], "reg_no": reg, "name": name,
                                  "reason": f"Class '{class_raw}' does not exist in the system"})
                continue

            # Section validation + verification
            section = normalize_section(section_raw)
            if not section:
                failed_count += 1
                error_log.append({"row": r["row"], "reg_no": reg, "name": name,
                                  "reason": "Section is empty or invalid"})
                continue
            if section_mode == "strict" and (class_name, section.lower()) not in existing_combos:
                failed_count += 1
                error_log.append({"row": r["row"], "reg_no": reg, "name": name,
                                  "reason": f"Section '{section}' does not exist for {class_name}"})
                continue

            existing = existing_by_login.get(reg)
            if existing:
                if dup_mode == "update":
                    if existing.role != "student":
                        failed_count += 1
                        error_log.append({"row": r["row"], "reg_no": reg, "name": name,
                                          "reason": f"Registration number belongs to a {existing.role} account"})
                        continue
                    existing.name = name
                    existing.father_name = father or existing.father_name
                    existing.class_name = class_name
                    existing.section = section
                    existing.is_active = True
                    updated_count += 1
                else:
                    # skip / create_new → leave existing records untouched
                    skipped_count += 1
                    error_log.append({"row": r["row"], "reg_no": reg, "name": name,
                                      "reason": "Registration number already exists — skipped"})
                continue

            # ── Create a new account ──
            if pw_mode == "custom" and custom_hash:
                password = custom_pw
                pw_hash = custom_hash
            elif pw_mode == "random":
                password = _random_password()
                pw_hash = hash_password(password)
            else:  # registration (default)
                password = reg
                pw_hash = hash_password(password)

            user = User(
                name=name,
                login_id=reg,
                father_name=father or None,
                password_hash=pw_hash,
                role="student",
                class_name=class_name,
                section=section,
                is_active=True,
                must_change_password=True,
            )
            new_users.append(user)
            created_count += 1
            credentials_rows.append({
                "reg_no": reg, "name": name, "father_name": father,
                "class_name": class_name, "section": section,
                "username": reg, "password": password,
            })

        if new_users:
            db.add_all(new_users)
            await db.flush()  # populate ids
            created_ids = [u.id for u in new_users]

        # Generate credentials Excel (created accounts only)
        has_credentials = False
        if credentials_rows:
            try:
                excel_bytes = generate_credentials_excel(credentials_rows, batch.filename)
                _save_credentials_file(batch_id, excel_bytes)
                has_credentials = True
            except Exception as exc:  # pragma: no cover - report but don't fail the import
                error_log.append({"row": 0, "reg_no": "", "name": "",
                                  "reason": f"Credentials file could not be generated: {exc}"})

        batch.created_count = created_count
        batch.updated_count = updated_count
        batch.skipped_count = skipped_count
        batch.failed_count = failed_count
        batch.error_log = error_log
        batch.created_user_ids = created_ids
        batch.has_credentials = has_credentials
        batch.status = "completed"
        batch.finished_at = _utcnow()
        await db.commit()

    except Exception as exc:
        await db.rollback()
        try:
            res = await db.execute(select(StudentImportBatch).where(StudentImportBatch.id == batch_id))
            b = res.scalar_one_or_none()
            if b:
                b.status = "failed"
                b.error_message = str(exc)[:1000]
                b.finished_at = _utcnow()
                await db.commit()
        except Exception:
            pass


# ─── CREDENTIALS FILE ─────────────────────────────────────────────────────────

import os

_IMPORTS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads", "imports"
)


def import_source_path(batch_id: str) -> str:
    """Path where the uploaded source .xlsx is staged for the worker to read."""
    return os.path.join(_IMPORTS_DIR, f"{batch_id}_source.xlsx")


def ensure_imports_dir() -> str:
    os.makedirs(_IMPORTS_DIR, exist_ok=True)
    return _IMPORTS_DIR


def credentials_path(batch_id: str) -> str:
    return os.path.join(_IMPORTS_DIR, f"{batch_id}_credentials.xlsx")


def _save_credentials_file(batch_id: str, data: bytes) -> None:
    os.makedirs(_IMPORTS_DIR, exist_ok=True)
    with open(credentials_path(batch_id), "wb") as f:
        f.write(data)


def generate_credentials_excel(rows: list[dict], title: str | None = None) -> bytes:
    """Build the credentials workbook: Reg #, Name, Father, Class, Username, Password."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Credentials"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:G1")
    ws["A1"] = "LSS Bot — Student Login Credentials"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = "CONFIDENTIAL — Share with students / parents only"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="FF0000")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Registration #", "Student Name", "Father Name", "Class", "Section", "Username", "Password"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row = 5
    for i, rec in enumerate(rows):
        values = [
            rec.get("reg_no", ""), rec.get("name", ""), rec.get("father_name", ""),
            rec.get("class_name", ""), rec.get("section", ""),
            rec.get("username", ""), rec.get("password", ""),
        ]
        for col, v in enumerate(values, 1):
            ws.cell(row=row, column=col, value=v).border = border
        if i % 2 == 1:
            light = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = light
        row += 1

    ws.cell(row=row + 1, column=1, value=f"Total accounts: {len(rows)}").font = Font(bold=True)

    widths = {"A": 16, "B": 26, "C": 26, "D": 12, "E": 10, "F": 16, "G": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
